import streamlit as st

import pandas as pd

import smtplib

from email.mime.text import MIMEText

from email.mime.multipart import MIMEMultipart

from email.mime.base import MIMEBase

from email import encoders

import time

import re

import json

import io

import urllib.parse

from datetime import datetime



# --- TRY TO LOAD RICH TEXT EDITOR (GMAIL STYLE EDITOR) ---

try:

    from streamlit_quill import st_quill

    QUILL_AVAILABLE = True

except ImportError:

    QUILL_AVAILABLE = False



# --- TRY TO LOAD AGGRID FOR GOOGLE SHEETS STYLE TABLES ---

try:

    from st_aggrid import AgGrid, GridOptionsBuilder, DataReturnMode, GridUpdateMode

    AGGRID_AVAILABLE = True

except ImportError:

    AGGRID_AVAILABLE = False



# --- TRY TO LOAD OPENPYXL FOR STYLED EXCEL EXPORTS ---

try:

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    from openpyxl.utils import get_column_letter

    OPENPYXL_AVAILABLE = True

except ImportError:

    OPENPYXL_AVAILABLE = False



# --- PAGE CONFIGURATION ---

st.set_page_config(page_title="Real Estate Tool Hub", layout="wide")



# --- REAL ESTATE BACKGROUND & GLASS UI ---

def set_real_estate_background():

    custom_css = """

    <style>

    /* 1. The Real Estate Background Image */

    .stApp, [data-testid="stAppViewContainer"] {

        background-image: url('https://images.unsplash.com/photo-1600596542815-ffad4c1539a9?ixlib=rb-4.0.3&auto=format&fit=crop&w=2000&q=80');

        background-size: cover;

        background-position: center;

        background-attachment: fixed;

        background-repeat: no-repeat;

    }

    

    /* 2. Frosted Glass Effect for Main Content Area */

    .block-container {

        background: rgba(255, 255, 255, 0.85); 

        backdrop-filter: blur(12px);

        -webkit-backdrop-filter: blur(12px);

        border-radius: 20px;

        padding: 3rem !important;

        margin-top: 2rem;

        margin-bottom: 2rem;

        box-shadow: 0 8px 32px 0 rgba(0, 0, 0, 0.3);

        border: 1px solid rgba(255, 255, 255, 0.5);

    }

    

    /* 3. Frosted Glass Effect for Sidebar */

    [data-testid="stSidebar"] {

        background-color: rgba(255, 255, 255, 0.80) !important;

        backdrop-filter: blur(12px);

        -webkit-backdrop-filter: blur(12px);

        border-right: 1px solid rgba(255, 255, 255, 0.5);

    }

    

    /* Make sure all text remains dark and readable even if Streamlit is in dark mode */

    h1, h2, h3, h4, p, label, .stRadio label, .stTabs [data-baseweb="tab-list"] button {

        color: #1f2937 !important; 

    }



    /* Red Primary Button Styling */

    div.stButton > button[kind="primary"], div.stDownloadButton > button[kind="primary"] {

        background-color: #ff4b4b !important;

        color: white !important;

        border: none !important;

        border-radius: 12px !important;

        font-weight: bold !important;

    }

    div.stButton > button[kind="primary"]:hover, div.stDownloadButton > button[kind="primary"]:hover {

        background-color: #e03838 !important;

    }

    </style>

    """

    st.markdown(custom_css, unsafe_allow_html=True)



set_real_estate_background()



# --- INITIALIZE SESSION STATE FOR PERSISTENT TEXT INPUTS & BIDIRECTIONAL SYNC ---

defaults = {

    "b_name": "VERONICA M. DIAZ",

    "b_age": 45,

    "b_cit": "US Citizen / Former Filipino",

    "b_fico": 650,

    "b_project": "SOUTH 2 RESIDENCES",

    "b_unit": "STH2/B030636",

    "b_tcp": 3991379.60,

    "b_ltv": 90.0,

    "b_lumpsum": 3592241.64,

    "exchange_rate": 58.00,

    "cfg_y1": 5,

    "cfg_r1": 7.25,

    "cfg_y2": 10,

    "cfg_r2": 7.75,

    "cfg_y3": 15,

    "cfg_r3": 8.00,

    "cfg_y4": 20,

    "cfg_r4": 8.50,

    "suite_bsal": 2500.0,

    "suite_bcomm": 0.0,

    "suite_bdis": 0.0,

    "suite_brent": 0.0,

    "suite_bbon1": 0.0,

    "suite_bbon2": 0.0,

    "suite_cbsal": 0.0,

    "suite_cbcomm": 0.0,

    "suite_cbdis": 0.0,

    "suite_cbbon1": 0.0,

    "suite_cbbon2": 0.0,

    "suite_d_re": 700.0,

    "suite_d_car": 0.0,

    "suite_d_cc": 0.0,

    "suite_d_oth": 0.0,

    "viber_phone": "+639922651295",

    "viber_balance": "₱3,991,379.60 which was past due since 6/12/2026",

    "comm_client": "VERONICA M. DIAZ",

    "comm_tcp": 8682079.95,

    "comm_rate": 1.30,

    "comm_tax": 10.00,

    "comm_split": 4.0

}



for key, val in defaults.items():

    if key not in st.session_state:

        st.session_state[key] = val



# --- ISOLATED MASTER DATA STORE FOR PERSISTENCE ---

if "store" not in st.session_state:

    st.session_state["store"] = defaults.copy()



store = st.session_state["store"]



# --- CALLBACK HELPERS ---

def save_field(key):

    store[key] = st.session_state[f"w_{key}"]



def save_tcp():

    store["b_tcp"] = st.session_state["w_b_tcp"]

    store["b_lumpsum"] = store["b_tcp"] * (store["b_ltv"] / 100.0)



def save_lumpsum():

    store["b_lumpsum"] = st.session_state["w_b_lumpsum"]

    if store["b_ltv"] > 0:

        store["b_tcp"] = store["b_lumpsum"] / (store["b_ltv"] / 100.0)



def save_ltv():

    store["b_ltv"] = st.session_state["w_b_ltv"]

    store["b_lumpsum"] = store["b_tcp"] * (store["b_ltv"] / 100.0)





# --- GLOBAL HELPER FUNCTIONS ---



@st.cache_data(ttl=300, show_spinner=False)

def load_cached_data(source):

    return pd.read_csv(source, encoding='utf-8-sig')



def clean_currency(value):

    if pd.isna(value):

        return 0.0

    cleaned_value = re.sub(r'[^\d.]', '', str(value))

    try:

        return float(cleaned_value) if cleaned_value else 0.0

    except ValueError:

        return 0.0



def extract_year(date_string):

    if pd.isna(date_string) or str(date_string).strip() == "":

        return None

    match = re.search(r'\b(20\d{2})\b', str(date_string))

    if match:

        return match.group(1)

    try:

        parsed_date = pd.to_datetime(date_string)

        return str(parsed_date.year)

    except:

        return None



def clean_html_whitespace(html_str):

    if not html_str:

        return ""

    return re.sub(r'>\s+<', '><', html_str.strip())



def format_plain_text_to_html(text):

    """Converts raw plain-text line breaks into HTML <br> tags while leaving HTML structures intact."""

    if not text:

        return ""

    has_html_tags = bool(re.search(r'<(p|br|div|ul|ol|li|table|tr|td|b|span|i)[\s/>]', text, re.IGNORECASE))

    if not has_html_tags:

        text = text.replace("\r\n", "\n").replace("\n", "<br>")

    else:

        text = text.replace("\r\n", "\n").replace("\n", "<br>")

    return text



def wrap_with_yellow_background(content_html):

    """Automatically wraps email content inside the light yellow container and attaches the standard signature image if not present."""

    if not content_html:

        return ""

    

    processed_content = format_plain_text_to_html(content_html)

    

    sign_off_html = '<br><br>Warm Regards,<br><br><img src="https://i.imgur.com/2b8igTi.png" width="350">'

    if 'https://i.imgur.com/2b8igTi.png' not in processed_content:

        processed_content += sign_off_html



    if processed_content.strip().startswith('<div style="font-family: Arial'):

        return processed_content

    return f'<div style="font-family: Arial, sans-serif; background-color: lightyellow; color: black; line-height: 1.5; font-size: 13px; padding: 15px;">{processed_content}</div>'



def calculate_monthly_amortization(principal, annual_rate_pct, years):

    if principal <= 0 or years <= 0:

        return 0.0

    monthly_rate = (annual_rate_pct / 100) / 12

    num_payments = years * 12

    if monthly_rate == 0:

        return principal / num_payments

    amortization = principal * (monthly_rate * (1 + monthly_rate)**num_payments) / ((1 + monthly_rate)**num_payments - 1)

    return amortization



def convert_df_to_styled_excel(df):

    output = io.BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:

        df.to_excel(writer, index=False, sheet_name='Filtered Report')

        worksheet = writer.sheets['Filtered Report']



        header_fill = PatternFill(start_color="0984E3", end_color="0984E3", fill_type="solid")

        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

        align_left = Alignment(horizontal="left", vertical="center")

        thin_border = Border(

            left=Side(style='thin', color='DDDDDD'),

            right=Side(style='thin', color='DDDDDD'),

            top=Side(style='thin', color='DDDDDD'),

            bottom=Side(style='thin', color='DDDDDD')

        )



        for cell in worksheet[1]:

            cell.fill = header_fill

            cell.font = header_font

            cell.alignment = align_left



        for col in worksheet.columns:

            max_len = 0

            col_letter = get_column_letter(col[0].column)

            for cell in col:

                if cell.row != 1:

                    cell.border = thin_border

                if cell.value is not None:

                    max_len = max(max_len, len(str(cell.value)))

            worksheet.column_dimensions[col_letter].width = max(max_len + 4, 12)



    return output.getvalue()





def generate_dashboard_data(input_url):

    df = load_cached_data(input_url).copy()

    df.columns = df.columns.str.strip()

    

    required_cols = ['BANK STATUS', 'LUMPSUM BALANCE', 'ENDORSEMENT DATE', 'RELEASED DATE', 'DEVELOPER']

    for col in required_cols:

        if col not in df.columns:

            st.error(f"Error: Missing '{col}' column in the dataset.")

            return None, None



    df['BANK STATUS'] = df['BANK STATUS'].astype(str).str.strip().str.upper()

    df['BANK STATUS'] = df['BANK STATUS'].replace({

        'CANCELLED | DENIED': 'CANCELLED',

        'UNDERREVIEW': 'UNDER REVIEW'

    })

    df['LUMPSUM BALANCE NUMERIC'] = df['LUMPSUM BALANCE'].apply(clean_currency)

    df['DEVELOPER'] = df['DEVELOPER'].astype(str).str.strip().str.upper()

    df['DEVELOPER'] = df['DEVELOPER'].replace({'NAN': 'UNKNOWN', '': 'UNKNOWN'})



    df['Endorsement_Year'] = df['ENDORSEMENT DATE'].apply(extract_year)

    

    if 'RELEASED DATE' in df.columns:

        df['Released_Year'] = df['RELEASED DATE'].apply(extract_year)

    else:

        df['Released_Year'] = None



    def determine_record_year(row):

        if row.get('BANK STATUS', '') == 'RELEASED':

            year = row.get('Released_Year')

            return year if pd.notna(year) and year else (row.get('Endorsement_Year') if pd.notna(row.get('Endorsement_Year')) and row.get('Endorsement_Year') else "Unknown")

        else:

            return row.get('Endorsement_Year') if pd.notna(row.get('Endorsement_Year')) and row.get('Endorsement_Year') else "Unknown"



    df['Record_Year'] = df.apply(determine_record_year, axis=1)



    target_statuses = ['APPROVED', 'ONGOING SUBMISSION', 'UNDER REVIEW', 'CANCELLED', 'RELEASED']

    filtered_df = df[df['BANK STATUS'].isin(target_statuses)].copy()



    records = filtered_df[['DEVELOPER', 'Record_Year', 'BANK STATUS', 'LUMPSUM BALANCE NUMERIC']].to_dict(orient='records')

    data_json = json.dumps(records)



    html_content = f"""

    <!DOCTYPE html>

    <html lang="en">

    <head>

        <meta charset="UTF-8">

        <meta name="viewport" content="width=device-width, initial-scale=1.0">

        <title>Interactive Leads Dashboard</title>

        <style>

            body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; margin: 0; padding: 40px; display: flex; flex-direction: column; align-items: center; background-image: url('https://images.unsplash.com/photo-1600596542815-ffad4c1539a9?ixlib=rb-4.0.3&auto=format&fit=crop&w=2000&q=80'); background-size: cover; background-position: center; background-attachment: fixed; color: #333; }}

            .dashboard-wrapper {{ background-color: rgba(255, 255, 255, 0.93); padding: 40px; border-radius: 12px; box-shadow: 0 10px 35px rgba(0, 0, 0, 0.25); display: flex; flex-direction: column; align-items: center; width: 100%; max-width: 1100px; backdrop-filter: blur(8px); }}

            h1 {{ color: #2c3e50; margin-top: 0; margin-bottom: 25px; text-align: center; text-transform: uppercase; letter-spacing: 1px; }}

            .control-panel {{ background-color: #ffffff; padding: 20px 30px; border-radius: 8px; box-shadow: 0 4px 15px rgba(0, 0, 0, 0.05); margin-bottom: 25px; display: flex; flex-wrap: wrap; gap: 30px; align-items: flex-start; width: 100%; box-sizing: border-box; border: 1px solid #eee; }}

            .control-group {{ display: flex; flex-direction: column; gap: 8px; }}

            .control-group label {{ font-weight: bold; color: #555; font-size: 0.95em; }}

            select {{ padding: 10px; font-size: 1em; border-radius: 5px; border: 1px solid #ccc; min-width: 220px; cursor: pointer; background-color: #f9f9f9; }}

            select:hover {{ border-color: #0984e3; }}

            .multi-select-dropdown {{ position: relative; min-width: 220px; }}

            .multi-select-dropdown summary {{ padding: 10px 15px; background-color: #f9f9f9; border: 1px solid #ccc; border-radius: 5px; cursor: pointer; font-size: 0.95em; list-style: none; display: flex; justify-between; align-items: center; }}

            .multi-select-dropdown summary::-webkit-details-marker {{ display: none; }}

            .multi-select-dropdown summary::after {{ content: '▼'; font-size: 0.8em; color: #555; }}

            .multi-select-dropdown[open] summary::after {{ content: '▲'; }}

            .multi-select-dropdown summary:hover {{ border-color: #0984e3; }}

            .multi-select-content {{ position: absolute; top: calc(100% + 5px); left: 0; width: 100%; min-width: 250px; background: #ffffff; border: 1px solid #ccc; border-radius: 5px; box-shadow: 0 4px 15px rgba(0, 0, 0, 0.15); z-index: 100; max-height: 250px; overflow-y: auto; display: flex; flex-direction: column; padding: 10px; gap: 5px; }}

            .multi-select-content label {{ display: flex; align-items: center; gap: 10px; cursor: pointer; font-size: 0.95em; padding: 8px; border-radius: 4px; font-weight: normal; }}

            .multi-select-content label:hover {{ background-color: #f1f1f1; }}

            .styled-table {{ border-collapse: collapse; font-size: 1.1em; width: 100%; background-color: #ffffff; border-radius: 8px 8px 0 0; overflow: hidden; box-shadow: 0 0 20px rgba(0, 0, 0, 0.08); border: 1px solid #eee; }}

            .styled-table thead tr {{ background-color: #0984e3; color: #ffffff; text-align: left; }}

            .styled-table th, .styled-table td {{ padding: 15px 20px; }}

            .styled-table tbody tr {{ border-bottom: 1px solid #dddddd; }}

            .styled-table tbody tr:nth-of-type(even) {{ background-color: #f9f9f9; }}

            .styled-table tbody tr:hover {{ background-color: #f1f1f1; }}

            .grand-total {{ border-bottom: 3px solid #0984e3 !important; font-weight: bold; background-color: #e3f2fd !important; color: #000; }}

            .footer-text {{ margin-top: 20px; font-size: 0.85em; color: #666; font-style: italic; }}

        </style>

    </head>

    <body>

        <div class="dashboard-wrapper">

            <h1>Interactive Leads Summary Dashboard</h1>

            <div class="control-panel">

                <div class="control-group">

                    <label for="devFilter">1. Filter by Developer</label>

                    <select id="devFilter" onchange="processDataFilters()">

                        <option value="ALL">All Developers</option>

                    </select>

                </div>

                <div class="control-group">

                    <label>2. Filter by Year</label>

                    <details class="multi-select-dropdown">

                        <summary>Select Years...</summary>

                        <div class="multi-select-content" id="yearCheckboxes"></div>

                    </details>

                </div>

                <div class="control-group">

                    <label>3. Filter by Status</label>

                    <details class="multi-select-dropdown">

                        <summary>Hide / Show Statuses...</summary>

                        <div class="multi-select-content" id="statusCheckboxes"></div>

                    </details>

                </div>

            </div>

            <table class="styled-table" id="reportTable">

                <thead>

                    <tr><th>Status</th><th>Count</th><th>Total Lumpsum Balance</th></tr>

                </thead>

                <tbody></tbody>

            </table>

            <p class="footer-text">Data securely exported from Google Sheets. Filters apply instantly.</p>

        </div>

        <script>

            const rawData = {data_json};

            let hiddenStatuses = new Set();

            let currentSummaryData = {{}};



            const formatCurrency = (amount) => '₱' + amount.toLocaleString('en-US', {{ minimumFractionDigits: 2, maximumFractionDigits: 2 }});



            function getStatusPriority(statusLabel) {{

                const label = statusLabel.toUpperCase();

                if (label.startsWith("RELEASED")) return 1;

                if (label.startsWith("APPROVED")) return 2;

                if (label.startsWith("UNDER REVIEW")) return 3;

                if (label.startsWith("ONGOING SUBMISSION")) return 4;

                if (label.startsWith("CANCELLED")) return 5;

                return 6; 

            }}



            document.addEventListener('click', function(event) {{

                const details = document.querySelectorAll('details.multi-select-dropdown');

                details.forEach(detail => {{

                    if (!detail.contains(event.target)) detail.removeAttribute('open');

                }});

            }});



            function initializeFilters() {{

                const devSet = new Set(); const yearSet = new Set();

                rawData.forEach(row => {{

                    if(row['DEVELOPER']) devSet.add(row['DEVELOPER']);

                    if(row['Record_Year']) yearSet.add(row['Record_Year']);

                }});

                const devSelect = document.getElementById('devFilter');

                Array.from(devSet).sort().forEach(dev => {{

                    let opt = document.createElement('option');

                    opt.value = dev; opt.textContent = dev; devSelect.appendChild(opt);

                }});

                const yearContainer = document.getElementById('yearCheckboxes');

                Array.from(yearSet).sort().forEach(year => {{

                    let label = document.createElement('label'); let cb = document.createElement('input');

                    cb.type = 'checkbox'; cb.value = year; cb.checked = true; cb.onchange = processDataFilters;

                    label.appendChild(cb); label.appendChild(document.createTextNode(' ' + year));

                    yearContainer.appendChild(label);

                }});

            }}



            function processDataFilters() {{

                const selectedDev = document.getElementById('devFilter').value;

                const checkedBoxes = document.querySelectorAll('#yearCheckboxes input:checked');

                const selectedYears = Array.from(checkedBoxes).map(cb => cb.value);



                const filteredData = rawData.filter(row => {{

                    const devMatch = (selectedDev === "ALL" || row['DEVELOPER'] === selectedDev);

                    const yearMatch = selectedYears.includes(row['Record_Year']);

                    return devMatch && yearMatch;

                }});



                currentSummaryData = {{}}; const generatedStatuses = new Set();

                filteredData.forEach(row => {{

                    let status = row['BANK STATUS'];

                    if (selectedYears.length > 1 && (status === 'CANCELLED' || status === 'RELEASED')) {{

                        status = status + ' (' + row['Record_Year'] + ')';

                    }}

                    generatedStatuses.add(status);

                    if (!currentSummaryData[status]) currentSummaryData[status] = {{ count: 0, balance: 0 }};

                    currentSummaryData[status].count += 1;

                    currentSummaryData[status].balance += row['LUMPSUM BALANCE NUMERIC'];

                }});



                const statusContainer = document.getElementById('statusCheckboxes');

                statusContainer.innerHTML = ''; 

                Array.from(generatedStatuses).sort((a, b) => getStatusPriority(a) - getStatusPriority(b) || a.localeCompare(b)).forEach(status => {{

                    let label = document.createElement('label'); let cb = document.createElement('input');

                    cb.type = 'checkbox'; cb.value = status; cb.checked = !hiddenStatuses.has(status);

                    cb.onchange = (e) => {{

                        if (e.target.checked) hiddenStatuses.delete(status);

                        else hiddenStatuses.add(status);

                        renderFinalTable(); 

                    }};

                    label.appendChild(cb); label.appendChild(document.createTextNode(' ' + status));

                    statusContainer.appendChild(label);

                }});

                renderFinalTable();

            }}



            function renderFinalTable() {{

                const tbody = document.querySelector('#reportTable tbody');

                tbody.innerHTML = ''; 

                const sortedStatuses = Object.keys(currentSummaryData).sort((a, b) => getStatusPriority(a) - getStatusPriority(b) || a.localeCompare(b));

                let grandTotalCount = 0; let grandTotalBalance = 0; let rowsAdded = 0;



                sortedStatuses.forEach(status => {{

                    if (!hiddenStatuses.has(status)) {{

                        const tr = document.createElement('tr');

                        tr.innerHTML = `<td>${{status}}</td><td>${{currentSummaryData[status].count}}</td><td>${{formatCurrency(currentSummaryData[status].balance)}}</td>`;

                        tbody.appendChild(tr);

                        grandTotalCount += currentSummaryData[status].count; grandTotalBalance += currentSummaryData[status].balance; rowsAdded++;

                    }}

                }});



                if (rowsAdded === 0) {{

                    tbody.innerHTML = '<tr><td colspan="3" style="text-align:center;">No records found for the selected filters.</td></tr>'; return;

                }}

                if (grandTotalCount > 0) {{

                    const totalTr = document.createElement('tr'); totalTr.className = 'grand-total';

                    totalTr.innerHTML = `<td>GRAND TOTAL</td><td>${{grandTotalCount}}</td><td>${{formatCurrency(grandTotalBalance)}}</td>`;

                    tbody.appendChild(totalTr);

                }}

            }}



            window.onload = () => {{ initializeFilters(); processDataFilters(); }};

        </script>

    </body>

    </html>

    """

    return html_content, filtered_df



# --- SCRIPT 2 HELPER FUNCTIONS (EMAILER) ---

def send_bulk_emails(sender, app_password, df, subject, body, attachments):

    smtp_server = "smtp.gmail.com"

    smtp_port = 587

    try:

        server = smtplib.SMTP(smtp_server, smtp_port)

        server.starttls() 

        server.login(sender, app_password)



        progress_text = "Sending emails..."

        my_bar = st.progress(0, text=progress_text)

        total_emails = len(df)

        current_email_num = 1



        for index, row in df.iterrows():

            raw_email_string = str(row['Email']).replace(';', ',').replace('/', ',')

            email_list = raw_email_string.split(',')

            

            valid_emails = [e.strip().replace(' ', '') for e in email_list if '@' in e.strip().replace(' ', '')]

            if not valid_emails:

                current_email_num += 1

                continue 

                

            final_to_address = ", ".join(valid_emails)

            

            msg = MIMEMultipart()

            msg['From'] = sender

            msg['To'] = final_to_address

            msg['Subject'] = subject

            msg['Disposition-Notification-To'] = sender 

            

            personalized_body = body

            

            try:

                due_date = pd.to_datetime(str(row['DueDate']).strip())

                today = pd.Timestamp.today().normalize()

                if due_date < today:

                    personalized_body = personalized_body.replace(

                        "is approaching its turnover date on", 

                        "has been past due since"

                    )

                    personalized_body = personalized_body.replace(

                        "which is due by", 

                        "which was past due since"

                    )

                    personalized_body = personalized_body.replace(

                        "which is due on", 

                        "which was past due since"

                    )

                else:

                    personalized_body = personalized_body.replace(

                        "which is due by", 

                        "which is due on"

                    )

            except Exception:

                pass

            

            personalized_body = personalized_body.replace("[Name]", str(row['Name']))

            personalized_body = personalized_body.replace("[Project]", str(row['Project']))

            personalized_body = personalized_body.replace("[Balance]", str(row['Balance']))

            personalized_body = personalized_body.replace("[DueDate]", str(row['DueDate']))

            personalized_body = personalized_body.replace("[UnitCode]", str(row['UnitCode']))

            personalized_body = personalized_body.replace("[ContractNumber]", str(row['ContractNumber']))

            personalized_body = personalized_body.replace("[EndorsementDate]", str(row['EndorsementDate']))

            

            final_email_html = wrap_with_yellow_background(personalized_body)

            msg.attach(MIMEText(final_email_html, 'html'))



            if attachments:

                for file in attachments:

                    part = MIMEBase("application", "octet-stream")

                    part.set_payload(file.getvalue()) 

                    encoders.encode_base64(part) 

                    part.add_header("Content-Disposition", f"attachment; filename={file.name}")

                    msg.attach(part)



            server.sendmail(sender, valid_emails, msg.as_string())

            

            progress_percentage = int((current_email_num / total_emails) * 100)

            if progress_percentage > 100: progress_percentage = 100

            my_bar.progress(progress_percentage, text=f"Sent to {row['Name']} ({final_to_address})")

            

            current_email_num += 1 

            time.sleep(1)



        server.quit()

        return True

    except Exception as e:

        st.error(f"Failed to send: {e}")

        return False



def send_single_email(sender, app_password, to_email, subject, body, attachments):

    smtp_server = "smtp.gmail.com"

    smtp_port = 587

    try:

        server = smtplib.SMTP(smtp_server, smtp_port)

        server.starttls() 

        server.login(sender, app_password)

        

        msg = MIMEMultipart()

        msg['From'] = sender

        msg['To'] = to_email

        msg['Subject'] = subject

        msg['Disposition-Notification-To'] = sender

        

        final_email_html = wrap_with_yellow_background(body)

        msg.attach(MIMEText(final_email_html, 'html'))

        

        if attachments:

            for file in attachments:

                part = MIMEBase("application", "octet-stream")

                part.set_payload(file.getvalue()) 

                encoders.encode_base64(part) 

                part.add_header("Content-Disposition", f"attachment; filename={file.name}")

                msg.attach(part)

                

        with st.spinner("Sending email..."):

            server.sendmail(sender, to_email, msg.as_string())

            

        server.quit()

        return True

    except Exception as e:

        st.error(f"Failed to send: {e}")

        return False





# --- SIDEBAR NAVIGATION ---

st.sidebar.title("Navigation Menu")

st.sidebar.write("Choose the tool you want to run:")

app_mode = st.sidebar.radio("Action", [

    "Dashboard Generator 📊", 

    "Emailer 🤖📨", 

    "Report Generator & Lookup 📑",

    "All-In-One Pre-Qual Suite 📑🧮",

    "Viber / WhatsApp Linker 💬",

    "Document Checklist Generator 📋",

    "Application Form Helper 📄",

    "Commission Tracker 💰",

    "Call & Follow-up Scheduler 📅"

])



# --- MODE 1: DASHBOARD GENERATOR ---

if app_mode == "Dashboard Generator 📊":

    st.title("Interactive Dashboard Generator 📊")

    st.write("Generate a standalone, interactive HTML dashboard directly from your live Google Sheets data.")

    

    default_url = "https://docs.google.com/spreadsheets/d/e/2PACX-1vTUt3Lfzo7sp01OFUnXhY_r4n5XjRgRRLVJqaI3dYzjNXPFv1TIYG6AO0UjCeamDRDPuSR4oe6Q99CY/pub?gid=859880315&single=true&output=csv"

    input_url = st.text_input("Google Sheet CSV URL:", value=default_url)

    

    if st.button("Generate HTML Dashboard", type="primary"):

        with st.spinner("Downloading data and baking HTML..."):

            html_output, export_df = generate_dashboard_data(input_url)

            

            if html_output:

                st.success("Dashboard successfully generated!")

                

                st.markdown("---")

                st.write("### 📥 Download Report Options")

                

                d_col1, d_col2, d_col3 = st.columns(3)

                

                with d_col1:

                    if OPENPYXL_AVAILABLE and export_df is not None:

                        excel_data = convert_df_to_styled_excel(export_df)

                        st.download_button(

                            label="📊 Download Formatted Excel (.xlsx)",

                            data=excel_data,

                            file_name="Leads_Summary_Report.xlsx",

                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",

                            type="primary"

                        )

                    else:

                        st.warning("Install `openpyxl` for Excel formatting")

                        

                with d_col2:

                    if export_df is not None:

                        csv_data = export_df.to_csv(index=False, encoding='utf-8-sig').encode('utf-8-sig')

                        st.download_button(

                            label="📄 Download Standard CSV (.csv)",

                            data=csv_data,

                            file_name="Leads_Summary_Report.csv",

                            mime="text/csv",

                            type="primary"

                        )

                        

                with d_col3:

                    st.download_button(

                        label="🌐 Download Interactive HTML (.html)",

                        data=html_output,

                        file_name="Interactive_Summary_Report.html",

                        mime="text/html",

                        type="primary"

                    )

                

                with st.expander("Preview Dashboard Below", expanded=True):

                    st.components.v1.html(html_output, height=800, scrolling=True)



# --- MODE 2: EMAILER ---

elif app_mode == "Emailer 🤖📨":

    st.title("Automated Emailer 🤖📨")



    raw_draft_action_required = """Hi Mr./Mrs. <b><span style="color: red;">[Name]</span></b>,<br>

I hope this email finds you well.<br>

I’m reaching out because SMDC has notified us that your unit at <b><span style="color: red;">[Project]</span></b>, Contract Number <b><span style="color: red;">[ContractNumber]</span></b>, Unit Code <b><span style="color: red;">[UnitCode]</span></b> is approaching its turnover date on <b><span style="color: red;">[DueDate]</span></b> with a lumpsum balance of <b><span style="color: red;">[Balance]</span></b>.<br>

I understand that managing property details from overseas can be incredibly time-consuming. However, we need to secure your bank financing now to ensure we hit the turnover deadline and avoid any late penalties or risk of unit forfeiture from the developer.<br>

Our goal is to make this as seamless as possible for you. We partner directly with major banks (BDO, China Bank, PNBLA, etc.) and can help you secure the lowest possible rates, even if you are based abroad or are a Dual Citizen.<br>

To keep your unit on track, please let me know how you would like to proceed:

<ul style="margin-top: 5px; margin-bottom: 10px; padding-left: 20px;"><li>Option 1: You are still interested in bank financing. (Reply to this email with your available date and time for a quick 10-minute call, including your time zone).</li><li>Option 2: You are already paying in cash or using SMDC In-House financing. (Just reply "Cash" or "In-House" so we can update your file and stop our follow-ups).</li></ul>

If you prefer to chat via Viber or WhatsApp, you can reach me directly at <b><span style="color: red;"><i>+639922651295</i></span></b>.<br>

I look forward to helping you secure bank financing for your investment."""



    raw_draft_bank_rates = """Hi <b><span style="color: red;">[Name]</span></b>,<br><br>

This is Kitt from Bahai Deals; I'm one of the loan advisors.<br><br>

This is a quick update regarding <b><span style="color: red;">[Project]</span></b>, with a <i><b>Contract Number <span style="color: red;">[ContractNumber]</span></b></i>, <i><b>Unit Code <span style="color: red;">[UnitCode]</span></b></i>.<br>

You currently have a remaining balance of <b><span style="color: red;">[Balance]</span></b>, which is due on <b><span style="color: red;">[DueDate]</span></b>.<br><br>

Please see the list of options below for bank financing.<br><br>

<i><b>Note:</b><br>

Rate is subject to change without prior notice.<br>

Prevailing rate to be applied at the time of availment / loan release date (This means that even if your loan was previously approved at a lower rate, it may be subject to adjustment depending on the current market rate / promo rate of the bank.<br><br>

About the fixing period:<br>

The fixing rate period means that your interest rate stays the same for a set number of years (for example, 1 year, 3 years, or 5 years). During this time, your monthly payments won’t change.<br>

After the fixing period ends, the loan becomes subject to yearly repricing. This means the bank will review and adjust your interest rate once every year, based on current market rates. As a result, your monthly payment may go up or down depending on the new rate.</i><br><br>

<b><span style="color: #0000FF;">Bank Interest rate and Loan Term:</span></b>

<ul style="margin-top: 8px; margin-bottom: 15px; padding-left: 20px;">

<li><i><b><span style="color: red;">China Bank</span></b></i> (age limit: 65 years old)<ul style="padding-left: 20px; margin-top: 4px;"><li>6.50% fixed p.a. for 1 Year</li><li>7.00% fixed p.a. for 3 Years</li><li>7.25% fixed p.a. For 5 Years</li><li>7.75% fixed p.a. for 10 years</li></ul></li>

<li><i><b><span style="color: red;">BPI</span></b></i> (age limit: 65 years old) - Applicable only for Filipino and Dual Citizen<ul style="padding-left: 20px; margin-top: 4px;"><li>6.50% fixed p.a. for 1 Year</li><li>7.00% fixed p.a. for 3 Years</li><li>7.25% fixed p.a. For 5 Years</li></ul></li>

<li><i><b><span style="color: red;">BDO</span></b></i> (age limit: 70 y/o)<br>&gt; For Dual Citizen and Filipino -<ul style="padding-left: 20px; margin-top: 4px;"><li>7.50% fixed for 3 years</li><li>8.00% fixed for 5 years</li><li>8.50% fixed for 10 years</li></ul>&gt; For US Citizen/Former Filipino<ul style="padding-left: 20px; margin-top: 4px;"><li>9.75% fixed for 2 years</li><li>10.25% fixed for 5 years</li></ul></li>

<li><i><b><span style="color: red;">RCBC</span></b></i> (age limit: 70 y/o) - Applicable only to Dual and Filipino Citizen<ul style="padding-left: 20px; margin-top: 4px;"><li>7.50% fixed for 3 years</li><li>8.00% fixed for 5 years</li></ul></li>

<li><i><b><span style="color: red;">PNB Los Angeles</span></b></i><br><br>

<div style="margin-left: 10px; margin-bottom: 15px;">

<table border="1" cellpadding="8" cellspacing="0" style="border-collapse: collapse; border: 1.5px solid #000000; width: 100%; max-width: 750px; font-size: 12px; background-color: #ffffff; color: #000000;">

<thead>

<tr style="background-color: #f2f2f2; font-weight: bold; text-align: left;">

<th style="border: 1.5px solid #000000; padding: 8px; width: 18%;">Loan Term</th>

<th style="border: 1.5px solid #000000; padding: 8px; width: 32%;">Interest Rate</th>

<th style="border: 1.5px solid #000000; padding: 8px; width: 50%;">Rate Type</th>

</tr>

</thead>

<tbody>

<tr>

<td style="border: 1.5px solid #000000; padding: 8px; text-align: center; vertical-align: top;"><b>5 years</b></td>

<td style="border: 1.5px solid #000000; padding: 8px; vertical-align: top;">7.25% per annum fixed for 5 years</td>

<td style="border: 1.5px solid #000000; padding: 8px; vertical-align: top;">Fixed</td>

</tr>

<tr>

<td style="border: 1.5px solid #000000; padding: 8px; text-align: center; vertical-align: top;"><b>10 years</b></td>

<td style="border: 1.5px solid #000000; padding: 8px; vertical-align: top;">8.00% per annum fixed for 10 years</td>

<td style="border: 1.5px solid #000000; padding: 8px; vertical-align: top;">Fixed</td>

</tr>

<tr>

<td style="border: 1.5px solid #000000; padding: 8px; text-align: center; vertical-align: top;"><b>15 years</b></td>

<td style="border: 1.5px solid #000000; padding: 8px; vertical-align: top;">Fixed at 8.00% per annum (first 10 years). Then floating interest rate or variable adjusted interest rate on the 11th – 15<sup>th</sup> year.</td>

<td style="border: 1.5px solid #000000; padding: 8px; vertical-align: top;">Fixed then Variable rate adjusted. The floating interest rate or variable interest rate will be determined based on the Wall Street Journal Prime Rate prevailing at that time, plus a spread of 3.0% or 3.5% (depending on the approval of the loan).</td>

</tr>

<tr>

<td style="border: 1.5px solid #000000; padding: 8px; text-align: center; vertical-align: top;"><b>20 years</b></td>

<td style="border: 1.5px solid #000000; padding: 8px; vertical-align: top;">Fixed at 8.00% per annum (first 10 years); Then floating interest rate or variable adjusted interest rate on the 11th – 20<sup>th</sup> year.</td>

<td style="border: 1.5px solid #000000; padding: 8px; vertical-align: top;">Fixed then Variable rate adjusted. The floating interest rate or variable interest rate will be determined based on the Wall Street Journal Prime Rate prevailing at that time, plus a spread of 3.0% or 3.5% (depending on the approval of the loan).</td>

</tr>

</tbody>

</table>

</div>

</li>

</ul>

<b><span style="color: #0000FF;">Application requirements</span></b><br>

<div style="margin-top: 8px;">

<b><span style="color: red;">1. BASIC DOCUMENTS</span></b>

<ul style="margin-top: 4px; margin-bottom: 12px; padding-left: 20px;"><li>Two valid IDs and passports.</li><li>Duly filled-out China Bank Application Form</li><li>Dual Citizenship documents (if applicable)</li><li>Marriage Certificate for married and CENOMAR for single</li></ul>

<b><span style="color: red;">2. INCOME DOCUMENTS (EMPLOYED)</span></b>

<ul style="margin-top: 4px; margin-bottom: 12px; padding-left: 20px;"><li>Copies of the latest pay stubs for the last 3 consecutive months.</li><li>Certificate of Employment indicating your position, date started and salary</li></ul>

<b><span style="color: red;">3. Copy of the Contract to Sell</span></b> (the developer and its project must be accredited)<br><br>

<b><span style="color: red;">4. ATTORNEY IN FACT</span></b> - should be based in the Philippines (e.g. Parents, siblings, mother-in-law, or father-in-law)

<ul style="margin-top: 4px; margin-bottom: 12px; padding-left: 20px;"><li>Contact Details (Phone number and Email Address)</li><li>TIN or SSS Number</li><li>2 Valid IDs</li><li>Latest Proof of Billing indicating current address</li></ul>

</div><br>

Please let me know if you have any questions or concerns. You can reach me at the phone numbers indicated below. The bank application form is also attached for your convenience."""



    raw_free_draft = ""



    draft_action_required = clean_html_whitespace(raw_draft_action_required)

    draft_bank_rates = clean_html_whitespace(raw_draft_bank_rates)

    draft_free = clean_html_whitespace(raw_free_draft)



    col1, col2 = st.columns(2)

    with col1:

        st.subheader("Sending Mode")

        sending_mode = st.radio("How do you want to send?", ["Bulk (Google Sheet/CSV)", "Individual (Single Email)"])

    

    with col2:

        st.subheader("Template Selection")

        template_choice = st.selectbox("Select Email Draft:", options=["Action Required", "Bank Rates", "Free Draft"], index=0)



    if template_choice == "Action Required":

        current_draft_text = draft_action_required

    elif template_choice == "Bank Rates":

        current_draft_text = draft_bank_rates

    else:

        current_draft_text = draft_free



    with st.form("email_form"):

        st.subheader("Account Credentials")

        sender_email = st.text_input("Your Email Address")

        password = st.text_input("App Password", type="password")

        

        if sending_mode == "Bulk (Google Sheet/CSV)":

            st.subheader("Client Data Source")

            st.caption("Choose ONE way to import your clients (leave the other blank):")

            spreadsheet_file = st.file_uploader("Option 1: Upload your .csv file", type=["csv"])

            st.write("**— OR —**")

            sheet_url = st.text_input("Option 2: Paste your Google Sheet Link")

            

            st.subheader("Filter by Endorsement Date (Optional)")

            target_date = st.text_input("Endorsement Date Filter (e.g., 1/14/2026)")

        else:

            st.subheader("Recipient Details")

            st.caption("Enter the single email address you want to send this to.")

            single_recipient = st.text_input("Recipient Email Address")

            st.info("Tip: Because you are sending an individual email, don't forget to edit any placeholders like `[Name]` in the editor box below before sending!")

            

        st.subheader("Email Subject")

        subject = st.text_input("Subject")



        st.markdown("---")

        st.subheader("Email Message Editor")

        

        body = st.text_area("Email Body (HTML supported)", value=current_draft_text, height=450)



        st.markdown("---")

        attach_col, send_col = st.columns([3, 1])

        

        with attach_col:

            uploaded_files = st.file_uploader("Choose files to attach (Optional)", accept_multiple_files=True)

            

        with send_col:

            st.write(" ")

            st.write(" ")

            submitted = st.form_submit_button("🚀 Send Email(s)", type="primary", use_container_width=True)



    if submitted:

        if not sender_email or not password or not subject or not body:

            st.warning("Please fill in your email, password, subject, and body before sending.")

        else:

            if sending_mode == "Individual (Single Email)":

                if not single_recipient:

                    st.warning("Please enter a Recipient Email Address.")

                else:

                    success = send_single_email(sender_email, password, single_recipient, subject, body, uploaded_files)

                    if success:

                        st.success(f"Email successfully sent to {single_recipient}!")

            

            else:

                if not spreadsheet_file and not sheet_url:

                    st.warning("Please EITHER upload a CSV file OR paste a Google Sheet link!")

                elif spreadsheet_file and sheet_url:

                    st.warning("Please only use one method! Either upload a file OR paste a link, but not both at the same time.")

                else:

                    try:

                        with st.spinner('Loading client data...'):

                            if spreadsheet_file:

                                df = load_cached_data(spreadsheet_file).copy()

                            else:

                                gid_part = ""

                                if "gid=" in sheet_url:

                                    gid_string = sheet_url[sheet_url.find("gid="):]

                                    gid_value = gid_string.split('&')[0]

                                    gid_part = "&" + gid_value



                                if "/edit" in sheet_url:

                                    csv_url = sheet_url.split("/edit")[0] + "/export?format=csv" + gid_part

                                else:

                                    csv_url = sheet_url

                                    

                                df = load_cached_data(csv_url).copy()

                            

                            df.columns = df.columns.str.strip()

                            df = df.rename(columns={

                                "PRINCIPAL BORROWERS NAME": "Name", "PROJECT": "Project", "LUMPSUM BALANCE": "Balance",

                                "LSB DUE DATE": "DueDate", "UNIT CODE": "UnitCode", "CONTRACT NUMBER": "ContractNumber", 

                                "ENDORSEMENT DATE": "EndorsementDate", "EMAIL ADDRESS": "Email", "BANK STATUS": "Status" 

                            })

                            

                            required_columns = ['Name', 'Email', 'Project', 'Balance', 'DueDate', 'UnitCode', 'ContractNumber', 'EndorsementDate', 'Status']

                            df = df.filter(items=required_columns)

                            

                        if not all(col in df.columns for col in required_columns):

                             st.error(f"Oops! Missing columns. Expected: {required_columns}. Found: {list(df.columns)}")

                        else:

                            df = df.dropna(subset=['Email']) 

                            df['Email'] = df['Email'].astype(str).str.replace(' ', '').str.strip(',;')

                            df = df[df['Email'].str.contains('@', na=False)] 

                            df = df.drop_duplicates(subset=['Email'])

                            

                            df['Status'] = df['Status'].astype(str).str.strip().str.upper()

                            

                            # Keep UNDERREVIEW without spaces as requested

                            forbidden_statuses = ['CANCELLED', 'CANCELLED | DENIED', 'DENIED', 'UNDERREVIEW', 'APPROVED', 'RELEASED']

                            df = df[~df['Status'].isin(forbidden_statuses)]

                            

                            if target_date: 

                                try:

                                    filter_date = pd.to_datetime(target_date.strip()).date()

                                    df['EndorsementDate_Temp'] = pd.to_datetime(df['EndorsementDate']).dt.date

                                    df = df[df['EndorsementDate_Temp'] == filter_date]

                                except Exception as e:

                                    st.warning("Oops! I couldn't understand that date format. Please try writing it exactly like 1/14/2026.")

                                    df = pd.DataFrame() 

                            

                            if df.empty:

                                st.error("No valid clients were found to email!")

                            else:

                                with st.spinner(f'Sending {len(df)} emails...'):

                                    success = send_bulk_emails(sender_email, password, df, subject, body, uploaded_files)

                                    

                                if success:

                                    st.success("All emails sent successfully!")

                    except Exception as e:

                        st.error(f"Could not read the data. Error details: {e}")



# --- MODE 3: REPORT GENERATOR & LOOKUP ---

elif app_mode == "Report Generator & Lookup 📑":

    st.title("Report Generator & Client Lookup 📑")

    st.write("Extract clients based on specific filters, or search the entire database for a single individual.")

    

    st.subheader("Client Data Source")

    col1, col2 = st.columns(2)

    with col1:

        spreadsheet_file = st.file_uploader("Upload .csv file", type=["csv"], key="report_csv")

    with col2:

        sheet_url = st.text_input("Or paste Google Sheet Link", key="report_url")

        

    if spreadsheet_file or sheet_url:

        try:

            with st.spinner("Loading and processing data..."):

                if spreadsheet_file:

                    df = load_cached_data(spreadsheet_file).copy()

                else:

                    gid_part = ""

                    if "gid=" in sheet_url:

                        gid_string = sheet_url[sheet_url.find("gid="):]

                        gid_value = gid_string.split('&')[0]

                        gid_part = "&" + gid_value



                    if "/edit" in sheet_url:

                        csv_url = sheet_url.split("/edit")[0] + "/export?format=csv" + gid_part

                    else:

                        csv_url = sheet_url

                        

                    df = load_cached_data(csv_url).copy()

                

                df.columns = df.columns.str.strip()

                

                st.markdown("---")

                

                tab1, tab2 = st.tabs(["📋 Filtered Status Report", "🔍 Master Client Lookup"])

                

                with tab1:

                    st.subheader("Interactive Data Explorer")

                    

                    if 'BANK STATUS' not in df.columns:

                        st.error("Error: Could not find the 'BANK STATUS' column in the uploaded data.")

                    else:

                        df['BANK STATUS'] = df['BANK STATUS'].fillna('')

                        df['BANK STATUS'] = df['BANK STATUS'].astype(str).str.strip().str.upper()

                        df['BANK STATUS'] = df['BANK STATUS'].replace({

                            'CANCELLED | DENIED': 'CANCELLED',

                            'UNDERREVIEW': 'UNDER REVIEW',

                            'NAN': '',

                            'NONE': ''

                        })

                        

                        if 'DEVELOPER' in df.columns:

                            df['DEVELOPER'] = df['DEVELOPER'].astype(str).str.strip().str.upper()

                            df['DEVELOPER'] = df['DEVELOPER'].replace({'NAN': '', 'NONE': ''})

                        

                        target_statuses = ['APPROVED', 'UNDER REVIEW', 'ONGOING SUBMISSION', 'CANCELLED', 'RELEASED', '']

                        

                        if 'ENDORSEMENT DATE' in df.columns:

                            df['Endorsement_Year'] = df['ENDORSEMENT DATE'].apply(extract_year)

                        else:

                            df['Endorsement_Year'] = None



                        def determine_report_year(row):

                            return row.get('Endorsement_Year') if pd.notna(row.get('Endorsement_Year')) and row.get('Endorsement_Year') else "Unknown"



                        df['Record_Year'] = df.apply(determine_report_year, axis=1)

                        

                        filtered_df = df[df['BANK STATUS'].isin(target_statuses)].copy()



                        if 'Endorsement_Year' in filtered_df.columns:

                            filtered_df = filtered_df.drop(columns=['Endorsement_Year'])

                        if 'Record_Year' in filtered_df.columns:

                            filtered_df = filtered_df.drop(columns=['Record_Year'])



                        if AGGRID_AVAILABLE:

                            st.info("💡 **Column Filters:** Click the menu icon next to any column name to see the checkbox list!")

                            

                            gb = GridOptionsBuilder.from_dataframe(filtered_df)

                            gb.configure_default_column(

                                filter='agSetColumnFilter',  

                                filterable=True, 

                                sortable=True,

                                resizable=True,

                                flex=1, 

                                minWidth=150,

                                suppressMenuHide=True,       

                                floatingFilter=False         

                            )

                            gridOptions = gb.build()

                            

                            grid_response = AgGrid(

                                filtered_df, 

                                gridOptions=gridOptions, 

                                enable_enterprise_modules=True,  

                                height=600,

                                theme="alpine",

                                data_return_mode=DataReturnMode.FILTERED_AND_SORTED,

                                update_mode=GridUpdateMode.FILTERING_CHANGED | GridUpdateMode.MODEL_CHANGED

                            )

                            

                            export_df = grid_response['data']

                            if isinstance(export_df, list):

                                export_df = pd.DataFrame(export_df)

                        else:

                            st.warning("⚠️ **Want filters inside the column names?** You need to install the AgGrid plugin. Open your terminal/command prompt and run: `pip install streamlit-aggrid`")

                            st.dataframe(filtered_df, use_container_width=True)

                            export_df = filtered_df

                            

                        st.markdown("---")

                        st.write(f"### 📥 Download Extracted Data ({len(export_df)} rows)")

                        

                        if not export_df.empty:

                            columns_to_exclude = st.multiselect(

                                "Select specific columns to EXCLUDE from your downloaded file:", 

                                options=export_df.columns.tolist(), 

                                default=[]

                            )

                            

                            final_export_df = export_df.drop(columns=columns_to_exclude)

                            

                            d_col1, d_col2 = st.columns(2)

                            

                            with d_col1:

                                if OPENPYXL_AVAILABLE:

                                    excel_data = convert_df_to_styled_excel(final_export_df)

                                    st.download_button(

                                        label="📊 Download Formatted Excel (.xlsx)",

                                        data=excel_data,

                                        file_name="Filtered_Client_Report.xlsx",

                                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",

                                        type="primary"

                                    )

                                else:

                                    st.warning("Install `openpyxl` for Excel formatting: `pip install openpyxl`")

                            

                            with d_col2:

                                csv_export = final_export_df.to_csv(index=False, encoding='utf-8-sig').encode('utf-8-sig')

                                st.download_button(

                                    label="📄 Download Standard CSV (.csv)",

                                    data=csv_export,

                                    file_name="Filtered_Client_Report.csv",

                                    mime="text/csv"

                                )

                        else:

                            st.warning("No rows selected or found. Change your filters to download data.")



                with tab2:

                    st.subheader("Search Entire Database")

                    st.caption("This search ignores the filters in the tab above. It scans every column in your raw sheet.")

                    search_query = st.text_input("Search (Type name, unit code, or email):", placeholder="e.g. John Doe or TWR1-0123")

                    

                    if search_query:

                        mask = df.apply(lambda row: row.astype(str).str.contains(search_query, case=False, na=False).any(), axis=1)

                        results_df = df[mask]

                        

                        if results_df.empty:

                            st.warning(f"No clients found matching '{search_query}'. Try checking for typos or searching by just their last name.")

                        else:

                            st.success(f"Found {len(results_df)} matching record(s)!")

                            st.dataframe(results_df, use_container_width=True)

                        

        except Exception as e:

            st.error(f"Error loading data: {e}")



# --- COMBINED TOOL: ALL-IN-ONE PRE-QUALIFICATION SUITE ---

elif app_mode == "All-In-One Pre-Qual Suite 📑🧮":

    st.title("All-In-One Pre-Qualification Suite 📑🧮")

    st.write("Enter the borrower's details once to calculate Mortgage Amortization, DTI Capacity, Partner Bank Age/Citizenship/FICO/DTI Eligibility, and export a complete assessment certificate.")



    # --- SECTION 1: UNIFIED BORROWER & PROPERTY INPUTS ---

    st.subheader("1. Borrower & Property Parameters")

    col1, col2, col3 = st.columns(3)

    

    with col1:

        st.text_input("Borrower Full Name:", value=store["b_name"], key="w_b_name", on_change=save_field, args=("b_name",))

        st.number_input("Current Age (Years):", min_value=18, max_value=80, value=store["b_age"], key="w_b_age", on_change=save_field, args=("b_age",))

        cit_options = ["Filipino", "Dual Citizen", "US Citizen / Former Filipino"]

        st.selectbox("Citizenship Status:", cit_options, index=cit_options.index(store["b_cit"]) if store["b_cit"] in cit_options else 0, key="w_b_cit", on_change=save_field, args=("b_cit",))

        st.number_input("FICO Score:", step=10, value=store["b_fico"], help="Required for PNB LA (Min. 650 score)", key="w_b_fico", on_change=save_field, args=("b_fico",))

        

    with col2:

        st.text_input("Project Name:", value=store["b_project"], key="w_b_project", on_change=save_field, args=("b_project",))

        st.text_input("Unit Code:", value=store["b_unit"], key="w_b_unit", on_change=save_field, args=("b_unit",))

        st.number_input("Total Contract Price (TCP ₱):", step=50000.0, format="%.2f", value=store["b_tcp"], key="w_b_tcp", on_change=save_tcp)

        st.number_input("Lumpsum Balance (₱):", step=50000.0, format="%.2f", value=store["b_lumpsum"], key="w_b_lumpsum", on_change=save_lumpsum, help="Auto-calculates TCP based on LTV %, or type directly if TCP is unknown.")



    with col3:

        st.number_input("Requested LTV Loan Rate (%):", step=1.0, value=store["b_ltv"], key="w_b_ltv", on_change=save_ltv)

        st.number_input("Exchange Rate (USD to PHP ₱):", step=0.10, value=store["exchange_rate"], key="w_exchange_rate", on_change=save_field, args=("exchange_rate",))



    # Fetch values from persistent store

    b_name = store["b_name"]

    b_age = store["b_age"]

    b_cit = store["b_cit"]

    b_fico = store["b_fico"]

    b_project = store["b_project"]

    b_unit = store["b_unit"]

    b_tcp = store["b_tcp"]

    b_lumpsum = store["b_lumpsum"]

    b_ltv = store["b_ltv"]

    exchange_rate = store["exchange_rate"]



    # Core Calculations

    peso_loan_amount = b_lumpsum

    dollar_loan_amount = peso_loan_amount / exchange_rate if exchange_rate > 0 else 0.0

    equity_amount_peso = b_tcp - peso_loan_amount



    st.markdown("---")



    # --- SECTION 2: MORTGAGE AMORTIZATION COMPARISON & TERM CONFIGURATION ---

    st.subheader("2. Mortgage Amortization Comparison")

    st.caption(f"Total Contract Price: **₱{b_tcp:,.2f}** | Lumpsum Loan Balance: **₱{peso_loan_amount:,.2f}** (${dollar_loan_amount:,.2f}) | Equity: **₱{equity_amount_peso:,.2f}** ({100-b_ltv:.0f}%)")



    # TERM INTEREST RATES CONFIGURATION SUBSECTION

    st.markdown("##### Term Interest Rates Configuration")

    years_options = list(range(1, 21))

    

    tc1, tc2, tc3, tc4 = st.columns(4)

    with tc1:

        st.selectbox("Option 1 Term (Years):", years_options, index=years_options.index(store["cfg_y1"]) if store["cfg_y1"] in years_options else 4, key="w_cfg_y1", on_change=save_field, args=("cfg_y1",))

        st.number_input("Option 1 Rate (% p.a.):", step=0.25, format="%.2f", value=store["cfg_r1"], key="w_cfg_r1", on_change=save_field, args=("cfg_r1",))

    with tc2:

        st.selectbox("Option 2 Term (Years):", years_options, index=years_options.index(store["cfg_y2"]) if store["cfg_y2"] in years_options else 9, key="w_cfg_y2", on_change=save_field, args=("cfg_y2",))

        st.number_input("Option 2 Rate (% p.a.):", step=0.25, format="%.2f", value=store["cfg_r2"], key="w_cfg_r2", on_change=save_field, args=("cfg_r2",))

    with tc3:

        st.selectbox("Option 3 Term (Years):", years_options, index=years_options.index(store["cfg_y3"]) if store["cfg_y3"] in years_options else 14, key="w_cfg_y3", on_change=save_field, args=("cfg_y3",))

        st.number_input("Option 3 Rate (% p.a.):", step=0.25, format="%.2f", value=store["cfg_r3"], key="w_cfg_r3", on_change=save_field, args=("cfg_r3",))

    with tc4:

        st.selectbox("Option 4 Term (Years):", years_options, index=years_options.index(store["cfg_y4"]) if store["cfg_y4"] in years_options else 19, key="w_cfg_y4", on_change=save_field, args=("cfg_y4",))

        st.number_input("Option 4 Rate (% p.a.):", step=0.25, format="%.2f", value=store["cfg_r4"], key="w_cfg_r4", on_change=save_field, args=("cfg_r4",))



    cfg_y1, cfg_r1 = store["cfg_y1"], store["cfg_r1"]

    cfg_y2, cfg_r2 = store["cfg_y2"], store["cfg_r2"]

    cfg_y3, cfg_r3 = store["cfg_y3"], store["cfg_r3"]

    cfg_y4, cfg_r4 = store["cfg_y4"], store["cfg_r4"]



    # Calculate payments for the configured options

    m_opt1_p = calculate_monthly_amortization(peso_loan_amount, cfg_r1, cfg_y1)

    m_opt2_p = calculate_monthly_amortization(peso_loan_amount, cfg_r2, cfg_y2)

    m_opt3_p = calculate_monthly_amortization(peso_loan_amount, cfg_r3, cfg_y3)

    m_opt4_p = calculate_monthly_amortization(peso_loan_amount, cfg_r4, cfg_y4)



    st.write(" ")

    mc1, mc2, mc3, mc4 = st.columns(4)

    mc1.metric(f"{cfg_y1}-Year Term", f"₱{m_opt1_p:,.2f} / mo", f"{cfg_r1:.2f}% p.a.")

    mc2.metric(f"{cfg_y2}-Year Term", f"₱{m_opt2_p:,.2f} / mo", f"{cfg_r2:.2f}% p.a.")

    mc3.metric(f"{cfg_y3}-Year Term", f"₱{m_opt3_p:,.2f} / mo", f"{cfg_r3:.2f}% p.a.")

    mc4.metric(f"{cfg_y4}-Year Term", f"₱{m_opt4_p:,.2f} / mo", f"{cfg_r4:.2f}% p.a.")



    st.write(" ")

    # EXPLICIT CLIENT CHOICE FOR PRIMARY DTI BENCHMARK

    opt_choices = [

        f"Option 1 ({cfg_y1} Yrs @ {cfg_r1:.2f}%)",

        f"Option 2 ({cfg_y2} Yrs @ {cfg_r2:.2f}%)",

        f"Option 3 ({cfg_y3} Yrs @ {cfg_r3:.2f}%)",

        f"Option 4 ({cfg_y4} Yrs @ {cfg_r4:.2f}%)"

    ]

    

    selected_target_opt = st.selectbox(

        "Select Target Benchmark Option for DTI & Bank Qualification Assessment:",

        options=opt_choices,

        index=store.get("selected_target_opt_idx", 2),

        key="w_selected_target_opt",

        on_change=lambda: store.update({"selected_target_opt_idx": opt_choices.index(st.session_state["w_selected_target_opt"])})

    )



    if "Option 1" in selected_target_opt:

        b_term, b_rate = cfg_y1, cfg_r1

    elif "Option 2" in selected_target_opt:

        b_term, b_rate = cfg_y2, cfg_r2

    elif "Option 3" in selected_target_opt:

        b_term, b_rate = cfg_y3, cfg_r3

    else:

        b_term, b_rate = cfg_y4, cfg_r4



    primary_amort_peso = calculate_monthly_amortization(peso_loan_amount, b_rate, b_term)

    primary_amort_usd = primary_amort_peso / exchange_rate if exchange_rate > 0 else 0.0



    st.markdown("---")



    # --- SECTION 3: DETAILED INCOME & DEBTS (PAYSTUB / EARNINGS USD $) ---

    st.subheader("3. Debt-to-Income (DTI) Credit Capacity Analysis")



    col_left, col_right = st.columns(2)



    with col_left:

        st.markdown("##### 1. Monthly Income (Paystub / Earnings USD $)")

        st.caption("Enter monthly income components in USD ($):")

        

        with st.expander("Primary Borrower Earnings", expanded=True):

            b_salary = st.number_input("Borrower Salary ($):", step=100.0, format="%.2f", value=store["suite_bsal"], key="w_suite_bsal", on_change=save_field, args=("suite_bsal",))

            b_commission = st.number_input("Borrower Commission ($):", step=100.0, format="%.2f", value=store["suite_bcomm"], key="w_suite_bcomm", on_change=save_field, args=("suite_bcomm",))

            b_disability = st.number_input("Borrower Disability Benefits ($):", step=100.0, format="%.2f", value=store["suite_bdis"], key="w_suite_bdis", on_change=save_field, args=("suite_bdis",))

            b_rent = st.number_input("Borrower Rent Income ($):", step=100.0, format="%.2f", value=store["suite_brent"], key="w_suite_brent", on_change=save_field, args=("suite_brent",))

            b_bonus1 = st.number_input("Borrower Bonus 1 ($):", step=100.0, format="%.2f", value=store["suite_bbon1"], key="w_suite_bbon1", on_change=save_field, args=("suite_bbon1",))

            b_bonus2 = st.number_input("Borrower Bonus 2 ($):", step=100.0, format="%.2f", value=store["suite_bbon2"], key="w_suite_bbon2", on_change=save_field, args=("suite_bbon2",))

            

            borrower_gmi_usd = b_salary + b_commission + b_disability + b_rent + b_bonus1 + b_bonus2

            st.markdown(f"**Borrower Gross Monthly Income (GMI): `${borrower_gmi_usd:,.2f}`**")



        with st.expander("Co-Borrower Earnings (Optional)", expanded=False):

            cb_salary = st.number_input("Co-Borrower Salary ($):", step=100.0, format="%.2f", value=store["suite_cbsal"], key="w_suite_cbsal", on_change=save_field, args=("suite_cbsal",))

            cb_commission = st.number_input("Co-Borrower Commission ($):", step=100.0, format="%.2f", value=store["suite_cbcomm"], key="w_suite_cbcomm", on_change=save_field, args=("suite_cbcomm",))

            cb_disability = st.number_input("Co-Borrower Disability Benefits ($):", step=100.0, format="%.2f", value=store["suite_cbdis"], key="w_suite_cbdis", on_change=save_field, args=("suite_cbdis",))

            cb_bonus1 = st.number_input("Co-Borrower Bonus 1 ($):", step=100.0, format="%.2f", value=store["suite_cbbon1"], key="w_suite_cbbon1", on_change=save_field, args=("suite_cbbon1",))

            cb_bonus2 = st.number_input("Co-Borrower Bonus 2 ($):", step=100.0, format="%.2f", value=store["suite_cbbon2"], key="w_suite_cbbon2", on_change=save_field, args=("suite_cbbon2",))

            

            coborrower_gmi_usd = cb_salary + cb_commission + cb_disability + cb_bonus1 + cb_bonus2

            st.markdown(f"**Co-Borrower Gross Monthly Income (GMI): `${coborrower_gmi_usd:,.2f}`**")



        total_gmi_usd = borrower_gmi_usd + coborrower_gmi_usd

        total_gmi_peso = total_gmi_usd * exchange_rate

        st.success(f"**Total Gross Monthly Income (GMI): `${total_gmi_usd:,.2f}` (`₱{total_gmi_peso:,.2f}`)**")



    with col_right:

        st.markdown("##### 2. Existing Monthly Debts / Loans (USD $)")

        st.caption("Enter current monthly debt payments in USD ($):")

        

        debt_real_estate = st.number_input("Existing Real Estate Loans ($):", step=50.0, format="%.2f", value=store["suite_d_re"], key="w_suite_d_re", on_change=save_field, args=("suite_d_re",))

        debt_car = st.number_input("Car Loan ($):", step=50.0, format="%.2f", value=store["suite_d_car"], key="w_suite_d_car", on_change=save_field, args=("suite_d_car",))

        debt_cc = st.number_input("Credit Card Monthly Payment ($):", step=50.0, format="%.2f", value=store["suite_d_cc"], key="w_suite_d_cc", on_change=save_field, args=("suite_d_cc",))

        debt_other = st.number_input("Other Existing Loans ($):", step=50.0, format="%.2f", value=store["suite_d_oth"], key="w_suite_d_oth", on_change=save_field, args=("suite_d_oth",))

        

        total_existing_debts_usd = debt_real_estate + debt_car + debt_cc + debt_other

        total_existing_debts_peso = total_existing_debts_usd * exchange_rate

        st.info(f"**Total Existing Monthly Debts: `${total_existing_debts_usd:,.2f}` (`₱{total_existing_debts_peso:,.2f}`)**")



        st.markdown("##### 3. Total Monthly Obligation Summary")

        total_monthly_obligation_usd = primary_amort_usd + total_existing_debts_usd

        total_monthly_obligation_peso = total_monthly_obligation_usd * exchange_rate



        st.write(f"• **New Loan Monthly Amortization:** `${primary_amort_usd:,.2f}` (`₱{primary_amort_peso:,.2f}`)")

        st.write(f"• **Existing Monthly Debts:** `${total_existing_debts_usd:,.2f}` (`₱{total_existing_debts_peso:,.2f}`)")

        st.metric("Total Monthly Obligation", f"${total_monthly_obligation_usd:,.2f}", f"₱{total_monthly_obligation_peso:,.2f}")



    # DTI Ratio Verdict

    if total_gmi_usd > 0:

        dti_ratio = (total_monthly_obligation_usd / total_gmi_usd) * 100.0

    else:

        dti_ratio = 0.0



    max_allowed_obligation_usd = total_gmi_usd * 0.40

    max_allowed_obligation_peso = max_allowed_obligation_usd * exchange_rate



    dc1, dc2, dc3 = st.columns(3)

    dc1.metric("Calculated DTI Ratio", f"{dti_ratio:.2f}%", help="Total Obligations / Gross Monthly Income")

    dc2.metric("Target Maximum DTI Limit", "40.00%", delta="Bank Benchmark")

    dc3.metric("Max Allowable Monthly Obligation (40%)", f"${max_allowed_obligation_usd:,.2f}", f"₱{max_allowed_obligation_peso:,.2f}")



    if dti_ratio <= 40.0 and dti_ratio > 0:

        st.success(f"✅ **QUALIFIED (PASS):** **{b_name}** capacity passes the 40% DTI benchmark!")

    else:

        st.error(f"❌ **OVER CAPACITY (FAIL):** **{b_name}** DTI ratio is **{dti_ratio:.2f}%** (Exceeds 40% threshold).")

        diff_usd = total_monthly_obligation_usd - max_allowed_obligation_usd

        st.info(f"💡 **Recommendation:** Needs co-borrower or monthly debt reduction of **`${diff_usd:,.2f}`/mo (`₱{diff_usd * exchange_rate:,.2f}`/mo)** to qualify.")



    st.markdown("---")



    # --- SECTION 4: PARTNER BANK ELIGIBILITY CHECKER ---

    st.subheader(f"4. Partner Bank Qualification Matrix (FICO: {b_fico})")



    banks_data = [

        {"Bank": "China Bank", "Max_Age": 65, "Allowed_Citizenship": ["Filipino", "Dual Citizen"], "Min_FICO": None, "Max_DTI": None},

        {"Bank": "BPI", "Max_Age": 65, "Allowed_Citizenship": ["Filipino", "Dual Citizen"], "Min_FICO": None, "Max_DTI": None},

        {"Bank": "BDO", "Max_Age": 70, "Allowed_Citizenship": ["Filipino", "Dual Citizen", "US Citizen / Former Filipino"], "Min_FICO": None, "Max_DTI": None},

        {"Bank": "RCBC", "Max_Age": 70, "Allowed_Citizenship": ["Filipino", "Dual Citizen"], "Min_FICO": None, "Max_DTI": None},

        {"Bank": "PNB Los Angeles", "Max_Age": None, "Allowed_Citizenship": ["Filipino", "Dual Citizen", "US Citizen / Former Filipino"], "Min_FICO": 650, "Max_DTI": 40.0}

    ]



    qualified_banks = []

    cols = st.columns(len(banks_data))

    for idx, b in enumerate(banks_data):

        bank_name = b["Bank"]

        max_a = b["Max_Age"]

        allowed_c = b["Allowed_Citizenship"]

        min_fico = b["Min_FICO"]

        max_dti = b["Max_DTI"]



        # Calculate exact Max Eligible Loan Term based on current borrower age

        if max_a is not None:

            max_term_eligible = max(0, max_a - b_age)

            max_term_eligible = min(20, max_term_eligible)

        else:

            max_term_eligible = 20



        age_ok = b_age <= max_a if max_a is not None else True

        cit_ok = b_cit in allowed_c

        fico_ok = (b_fico >= min_fico) if min_fico is not None else True

        dti_ok = (dti_ratio <= max_dti) if max_dti is not None else True



        with cols[idx]:

            max_age_str = f"{max_a} y/o" if max_a is not None else "No Age Limit"

            term_str = f"Max Term: **{max_term_eligible} Yrs**" if max_term_eligible > 0 else "Max Term: **0 Yrs (Over Age)**"

            

            if age_ok and cit_ok and fico_ok and dti_ok and max_term_eligible > 0:

                st.success(f"✅ **{bank_name}**")

                st.caption(f"Max Age: {max_age_str} | {term_str}")

                if max_a is not None:

                    st.write(f"• **Age Rule:** PASS (Current Age {b_age} ≤ {max_a})")

                else:

                    st.write("• **Age Rule:** PASS")

                st.write("• **Citizenship:** PASS")

                if min_fico is not None:

                    st.write(f"• **FICO Rule:** PASS (≥ {min_fico})")

                if max_dti is not None:

                    st.write(f"• **DTI Rule:** PASS (≤ {max_dti:.0f}%)")

                qualified_banks.append(f"{bank_name} (Up to {max_term_eligible}-Yr Max Term)")

            else:

                st.error(f"❌ **{bank_name}**")

                st.caption(f"Max Age: {max_age_str} | {term_str}")

                if not age_ok or max_term_eligible <= 0:

                    st.write(f"• **Age Rule:** FAIL (Current Age {b_age} > {max_a})")

                else:

                    st.write(f"• **Age Rule:** PASS (Current Age {b_age} ≤ {max_a})")



                if not cit_ok:

                    st.write("• **Citizenship:** FAIL")

                else:

                    st.write("• **Citizenship:** PASS")



                if min_fico is not None:

                    if not fico_ok:

                        st.write(f"• **FICO Rule:** FAIL (< {min_fico})")

                    else:

                        st.write(f"• **FICO Rule:** PASS (≥ {min_fico})")



                if max_dti is not None:

                    if not dti_ok:

                        st.write(f"• **DTI Rule:** FAIL (> {max_dti:.0f}%)")

                    else:

                        st.write(f"• **DTI Rule:** PASS (≤ {max_dti:.0f}%)")



    st.markdown("---")



    # --- SECTION 5: PRE-QUALIFICATION CERTIFICATE GENERATOR ---

    st.subheader("5. Pre-Qualification Summary Certificate")



    bank_list_str = "\n".join([f"• {bk}" for bk in qualified_banks]) if qualified_banks else "• No matching partner banks found based on current criteria."



    prequal_cert = f"""======================================================================

              TALINO GROUP - LOAN PRE-QUALIFICATION ASSESSMENT

======================================================================

DATE GENERATED : {datetime.now().strftime('%B %d, %Y')}

ADVISOR        : Kitt (Bahai Deals | Talino Group)



[1] BORROWER PROFILE:

• Full Name           : {b_name}

• Current Age / Citizenship : {b_age} y/o | {b_cit}

• FICO Score          : {b_fico}

• Gross Monthly Income: ${total_gmi_usd:,.2f} (₱{total_gmi_peso:,.2f})

• Existing Obligations: ${total_existing_debts_usd:,.2f} (₱{total_existing_debts_peso:,.2f})



[2] PROPERTY & LOAN PARAMETERS:

• Project & Unit      : {b_project} ({b_unit})

• Total Contract Price: ₱{b_tcp:,.2f}

• Lumpsum Balance / Loan Amount Requested: ₱{peso_loan_amount:,.2f} (${dollar_loan_amount:,.2f})

• Downpayment Equity  : ₱{equity_amount_peso:,.2f} ({100-b_ltv:.0f}%)

• Proposed Term & Rate: {b_term} Years @ {b_rate:.2f}% p.a.

• Estimated Amortization: ₱{primary_amort_peso:,.2f} (${primary_amort_usd:,.2f}) / month



[3] CREDIT CAPACITY & DTI ASSESSMENT:

• Total Monthly Obligation : ${total_monthly_obligation_usd:,.2f} (₱{total_monthly_obligation_peso:,.2f})

• Debt-to-Income (DTI) Ratio: {dti_ratio:.2f}%

• DTI CAPACITY VERDICT      : {'QUALIFIED (PASS ≤ 40%)' if dti_ratio <= 40 else 'REQUIRES CO-BORROWER / DEBT REDUCTION (> 40%)'}



[4] QUALIFIED PARTNER BANKS FOR SUBMISSION (WITH ELIGIBLE MAX TERMS):

{bank_list_str}

======================================================================="""



    st.code(prequal_cert, language="text")

    st.download_button("📥 Download Full Pre-Qual Certificate (.txt)", data=prequal_cert, file_name=f"PreQual_Certificate_{b_name.replace(' ', '_')}.txt", mime="text/plain")



    st.markdown("---")



    # --- SECTION 6: QUICK COPY MESSAGE FOR VIBER / WHATSAPP / EMAIL ---

    st.subheader("📋 Quick Copy Message for Viber / WhatsApp / Email")



    cb_max_term = max(0, 65 - b_age)

    cb_max_term = min(20, cb_max_term)

    bdo_max_term = max(0, 70 - b_age)

    bdo_max_term = min(20, bdo_max_term)



    quick_text = f"""Hi {b_name}, here is your estimated monthly payment breakdown for a loan balance of ₱{peso_loan_amount:,.2f}:



• {cfg_y1}-Year Term ({cfg_r1:.2f}% p.a.): approx. ₱{m_opt1_p:,.2f} / month

• {cfg_y2}-Year Term ({cfg_r2:.2f}% p.a.): approx. ₱{m_opt2_p:,.2f} / month

• {cfg_y3}-Year Term ({cfg_r3:.2f}% p.a.): approx. ₱{m_opt3_p:,.2f} / month

• {cfg_y4}-Year Term ({cfg_r4:.2f}% p.a.): approx. ₱{m_opt4_p:,.2f} / month



Existing Obligations & Debt-To-Income Capacity:

• New Loan Monthly Amortization: ${primary_amort_usd:,.2f} (₱{primary_amort_peso:,.2f}) @ {b_rate:.2f}% p.a. ({b_term}-Year Term)

• Existing Monthly Debts: ${total_existing_debts_usd:,.2f} (₱{total_existing_debts_peso:,.2f})

• Total Combined Monthly Obligation: ${total_monthly_obligation_usd:,.2f} (₱{total_monthly_obligation_peso:,.2f})

• Calculated DTI Ratio: {dti_ratio:.2f}% (Bank Benchmark: ≤ 40%)



Eligible Max Terms Based on Current Age ({b_age} y/o):

• China Bank / BPI: Max {cb_max_term} Years

• BDO / RCBC: Max {bdo_max_term} Years

• PNB Los Angeles: Max 20 Years



Rates and terms are subject to bank approval. Let me know which option fits your monthly target best!"""



    st.code(quick_text, language="text")



# --- MODE 9: VIBER & WHATSAPP LINK GENERATOR WITH AUTO-FILL FROM GOOGLE SHEET ---

elif app_mode == "Viber / WhatsApp Linker 💬":

    st.title("Viber & WhatsApp Instant Messaging Linker 💬")

    st.write("Generate one-click direct chat links to contact clients instantly on messaging apps with personalized text pre-loaded.")



    default_sheet = "https://docs.google.com/spreadsheets/d/1TVrNsmb_J5RtMglqVNCtPcMD7vJKRiL9rGzoSGT2QRI/edit?gid=859880315#gid=859880315"

    

    st.subheader("🔍 Auto-Fill Client Details from Google Sheet")

    sheet_url_input = st.text_input("Google Sheet Data Link:", value=default_sheet)

    

    if sheet_url_input:

        try:

            gid_part = ""

            if "gid=" in sheet_url_input:

                gid_string = sheet_url_input[sheet_url_input.find("gid="):]

                gid_value = gid_string.split('&')[0]

                gid_part = "&" + gid_value



            if "/edit" in sheet_url_input:

                csv_url = sheet_url_input.split("/edit")[0] + "/export?format=csv" + gid_part

            else:

                csv_url = sheet_url_input



            client_df = load_cached_data(csv_url).copy()

            client_df.columns = client_df.columns.str.strip()



            name_col = None

            for col in ['PRINCIPAL BORROWERS NAME', 'NAME', 'BORROWER', 'CLIENT NAME']:

                if col in client_df.columns:

                    name_col = col

                    break

            

            if name_col:

                client_names = client_df[name_col].dropna().unique().tolist()

                selected_client = st.selectbox("Select / Search Client Name to Auto-Fill:", options=["-- Select a Client --"] + client_names)

                

                if selected_client != "-- Select a Client --":

                    row = client_df[client_df[name_col] == selected_client].iloc[0]

                    

                    store["b_name"] = str(row[name_col])

                    

                    proj = str(row.get('PROJECT', '')).strip()

                    unit = str(row.get('UNIT CODE', '')).strip()

                    if proj and unit:

                        store["b_unit"] = f"{unit} ({proj})"

                    elif proj:

                        store["b_unit"] = proj

                    elif unit:

                        store["b_unit"] = unit

                        

                    bal = str(row.get('LUMPSUM BALANCE', '')).strip()

                    due = str(row.get('LSB DUE DATE', '')).strip()

                    

                    due_text_phrase = f"due on {due}"

                    if due:

                        try:

                            parsed_due = pd.to_datetime(due)

                            today = pd.Timestamp.today().normalize()

                            if parsed_due < today:

                                due_text_phrase = f"which was past due since {due}"

                        except Exception:

                            pass

                            

                    if bal and due:

                        store["viber_balance"] = f"{bal} {due_text_phrase}"

                    elif bal:

                        store["viber_balance"] = bal

                        

                    phone_col = None

                    for pc in ['MOBILE NUMBER', 'CONTACT NUMBER', 'PHONE NUMBER', 'PHONE', 'MOBILE', 'CONTACT NO']:

                        if pc in client_df.columns:

                            phone_col = pc

                            break

                    if phone_col and pd.notna(row.get(phone_col)):

                        raw_p = str(row.get(phone_col)).strip()

                        if not raw_p.startswith('+') and not raw_p.startswith('63') and len(raw_p) >= 10:

                            raw_p = "+63" + raw_p.lstrip('0')

                        store["viber_phone"] = raw_p



        except Exception as e:

            st.warning(f"Could not automatically parse Google Sheet: {e}")



    st.markdown("---")

    st.subheader("Client Details & Message Options")



    col1, col2 = st.columns(2)

    with col1:

        client_name = st.text_input("Client Name:", value=store["b_name"], key="w_v_b_name", on_change=save_field, args=("b_name",))

        phone_number = st.text_input("Mobile / Phone Number (Include Country Code, e.g. +639922651295):", value=store["viber_phone"], key="w_viber_phone", on_change=save_field, args=("viber_phone",))

    with col2:

        unit_details = st.text_input("Unit Code & Project:", value=store["b_unit"], key="w_v_b_unit", on_change=save_field, args=("b_unit",))

        balance_due = st.text_input("Remaining Balance & Due Date Status:", value=store["viber_balance"], key="w_viber_balance", on_change=save_field, args=("viber_balance",))



    msg_type = st.selectbox("Select Message Template:", [

        "Turnover & Bank Financing Follow-up",

        "Document Request Reminder",

        "Bank Approval Good News Update"

    ])



    formatted_bal_due = balance_due

    if "due on " in formatted_bal_due.lower():

        try:

            date_match = re.search(r'due on\s+([\d/-]+)', formatted_bal_due, re.IGNORECASE)

            if date_match:

                extracted_date_str = date_match.group(1)

                parsed_due = pd.to_datetime(extracted_date_str)

                today = pd.Timestamp.today().normalize()

                if parsed_due < today:

                    formatted_bal_due = formatted_bal_due.replace(date_match.group(0), f"which was past due since {extracted_date_str}")

        except Exception:

            pass



    if msg_type == "Turnover & Bank Financing Follow-up":

        message_body = f"Hi {client_name}, this is Kitt from Bahai Deals| Talino Group! Just following up regarding your unit at {unit_details}. You have a balance of {formatted_bal_due}. Let me know if you would like us to secure bank financing for you now to avoid penalties -- No worries, our services are absolutely free and no hidden fees. Let me know your available time. I can walk you through prequalification process. Thank you."

    elif msg_type == "Document Request Reminder":

        message_body = f"Hi {client_name}, this is Kitt from Bahai Deals| Talino Group. To move forward with your bank loan application for {unit_details}, could you please send over your latest pay stubs and valid IDs? Thank you!"

    else:

        message_body = f"Hi {client_name}, great news! We have an update regarding your bank financing approval for {unit_details}. Please let me know when you're available for a quick 5-minute call. Thanks!"



    custom_msg = st.text_area("Personalize Your Message:", value=message_body, height=140)



    clean_phone = re.sub(r'[^\d]', '', phone_number)

    encoded_text = urllib.parse.quote(custom_msg)



    wa_link = f"whatsapp://send?phone={clean_phone}&text={encoded_text}"

    viber_link = f"viber://chat?number=%2B{clean_phone}"



    st.markdown("---")

    st.subheader("📲 Launch Native Desktop/Mobile App Directly")



    l_col1, l_col2 = st.columns(2)

    with l_col1:

        st.markdown(f'<a href="{wa_link}" style="display:inline-block; background-color:#25D366; color:white; padding:12px 24px; text-decoration:none; border-radius:8px; font-weight:bold; width:100%; text-align:center;">💬 Open WhatsApp App Directly</a>', unsafe_allow_html=True)

    with l_col2:

        st.markdown(f'<a href="{viber_link}" style="display:inline-block; background-color:#7360F2; color:white; padding:12px 24px; text-decoration:none; border-radius:8px; font-weight:bold; width:100%; text-align:center;">📱 Open Viber App Directly</a>', unsafe_allow_html=True)



# --- MODE 10: DOCUMENT CHECKLIST GENERATOR ---

elif app_mode == "Document Checklist Generator 📋":

    st.title("Document Checklist & Missing Docs Generator 📋")

    st.write("Track received items and generate a clear list of remaining missing requirements to send to your client.")



    c_col1, c_col2 = st.columns(2)

    with c_col1:

        client_name = st.text_input("Client Name:", value=store["b_name"], key="w_chk_b_name", on_change=save_field, args=("b_name",))

    with c_col2:

        target_bank = st.selectbox("Target Bank:", ["China Bank", "BPI", "BDO", "RCBC", "PNB Los Angeles", "General Bank Submission"])



    st.subheader("1. Basic Documents")

    b1 = st.checkbox("Two valid IDs and Passports", value=True)

    b2 = st.checkbox(f"Duly filled-out {target_bank} Application Form", value=False)

    b3 = st.checkbox("Dual Citizenship Documents (if applicable)", value=False)

    b4 = st.checkbox("Marriage Certificate (if married) or CENOMAR (if single)", value=False)



    st.subheader("2. Income Documents (Employed / OFW)")

    i1 = st.checkbox("Copies of latest pay stubs for the last 3 consecutive months", value=False)

    i2 = st.checkbox("Certificate of Employment (COE) indicating position, start date, and salary", value=False)



    st.subheader("3. Contract & Property Documents")

    p1 = st.checkbox("Copy of Contract to Sell (CTS)", value=True)



    st.subheader("4. Attorney-In-Fact (SPA) Requirements (If Overseas)")

    a1 = st.checkbox("SPA Contact Details (Phone number & Email)", value=False)

    a2 = st.checkbox("SPA TIN or SSS Number", value=False)

    a3 = st.checkbox("2 Valid IDs of Attorney-In-Fact", value=False)

    a4 = st.checkbox("Latest Proof of Billing (showing current address)", value=False)



    missing_docs = []

    if not b1: missing_docs.append("Two valid IDs and Passports")

    if not b2: missing_docs.append(f"Duly filled-out {target_bank} Application Form")

    if not b3: missing_docs.append("Dual Citizenship Documents (if applicable)")

    if not b4: missing_docs.append("Marriage Certificate (if married) or CENOMAR (if single)")

    if not i1: missing_docs.append("Copies of latest pay stubs for the last 3 consecutive months")

    if not i2: missing_docs.append("Certificate of Employment (COE) indicating position, start date, and salary")

    if not p1: missing_docs.append("Copy of Contract to Sell (CTS)")

    if not a1: missing_docs.append("SPA Contact Details (Phone number & Email)")

    if not a2: missing_docs.append("SPA TIN or SSS Number")

    if not a3: missing_docs.append("2 Valid IDs of Attorney-In-Fact")

    if not a4: missing_docs.append("Latest Proof of Billing indicating current address")



    st.markdown("---")

    st.subheader("📝 Generated Follow-Up Message")



    if not missing_docs:

        summary_msg = f"Hi {client_name}, good news! You have submitted all the necessary documents for your {target_bank} loan application. We are now processing your endorsement."

    else:

        doc_list_str = "\n".join([f"• {doc}" for doc in missing_docs])

        summary_msg = f"""Hi {client_name},



Hope you are doing well! To proceed with your bank loan application for {target_bank}, we are still requesting the following remaining requirements:



{doc_list_str}



Please email or reply with these documents at your earliest convenience so we can keep your financing on schedule. Thank you!"""



    st.text_area("Copy and paste to email or chat:", value=summary_msg, height=250)



# --- MODE 11: APPLICATION FORM HELPER ---

elif app_mode == "Application Form Helper 📄":

    st.title("Application Form Data Helper 📄")

    st.write("Extract client information into a formatted reference summary to easily paste into official bank PDF application forms.")



    col1, col2 = st.columns(2)

    with col1:

        borrower_name = st.text_input("Principal Borrower Name:", value=store["b_name"], key="w_app_b_name", on_change=save_field, args=("b_name",))

        civil_status = st.selectbox("Civil Status:", ["Single", "Married", "Widowed", "Separated"])

        email_addr = st.text_input("Email Address:", value="juan.delacruz@example.com")

        contact_no = st.text_input("Contact Number:", value="+639171234567")

    with col2:

        project_name = st.text_input("Project Name:", value=store["b_project"], key="w_app_b_project", on_change=save_field, args=("b_project",))

        unit_code = st.text_input("Unit Code:", value=store["b_unit"], key="w_app_b_unit", on_change=save_field, args=("b_unit",))

        contract_no = st.text_input("Contract Number:", value="12345678")

        loan_amount = st.text_input("Requested Loan Amount (₱):", value=f"₱{store['b_lumpsum']:,.2f}")



    st.markdown("---")

    st.subheader("📋 Bank Application Data Sheet Summary")



    data_summary = f"""==================================================

        BANK FINANCING BORROWER PROFILE

==================================================

BORROWER INFORMATION:

• Full Name: {borrower_name}

• Civil Status: {civil_status}

• Email: {email_addr}

• Contact Number: {contact_no}



PROPERTY & FINANCING DETAILS:

• Developer/Project: {project_name}

• Unit Code: {unit_code}

• Contract Number: {contract_no}

• Requested Loan Amount: {loan_amount}

=================================================="""



    st.code(data_summary, language="text")

    st.download_button("📥 Download Summary as .txt File", data=data_summary, file_name=f"Borrower_Profile_{borrower_name.replace(' ', '_')}.txt", mime="text/plain")



# --- MODE 12: COMMISSION & RELEASE TRACKER ---

elif app_mode == "Commission Tracker 💰":

    st.title("Commission & Release Tracker 💰")

    st.write("Track sales commissions, calculate withholding tax deductions, advisor splits, and project expected release dates.")



    col1, col2 = st.columns(2)

    with col1:

        client_name = st.text_input("Client Name:", value=store["comm_client"], key="w_comm_client", on_change=save_field, args=("comm_client",))

        tcp_amount = st.number_input("Total Contract Price (TCP ₱):", step=10000.0, format="%.2f", value=store["comm_tcp"], key="w_comm_tcp", on_change=save_field, args=("comm_tcp",))

        comm_pct = st.number_input("Commission Rate (%):", step=0.05, format="%.2f", value=store["comm_rate"], key="w_comm_rate", on_change=save_field, args=("comm_rate",))

        tax_pct = st.number_input("Withholding Tax Rate (%):", step=1.0, format="%.2f", value=store["comm_tax"], help="Tax percentage deducted from gross commission.", key="w_comm_tax", on_change=save_field, args=("comm_tax",))

        advisor_split_pct = st.number_input("Advisor Split Share (%):", step=0.5, format="%.2f", value=store["comm_split"], key="w_comm_split", on_change=save_field, args=("comm_split",))



    with col2:

        project_name = st.text_input("Project Name:", value=store["b_project"], key="w_comm_b_project", on_change=save_field, args=("b_project",))

        milestone = st.selectbox("Release Milestone:", [

            "Upon 10% Downpayment Paid",

            "Upon CTS Signing & Approval",

            "Upon Bank Loan Release / Takeout",

            "50% Initial Release / 50% Final Release"

        ])

        expected_date = st.date_input("Target Release Date:")

        status = st.selectbox("Payout Status:", ["Pending", "Processing", "Released", "On Hold"])



    # Calculations with Tax Deduction

    gross_commission = tcp_amount * (comm_pct / 100.0)

    tax_amount = gross_commission * (tax_pct / 100.0)

    net_effective_comm_rate = comm_pct * (1.0 - (tax_pct / 100.0))

    net_commission_after_tax = gross_commission - tax_amount

    

    advisor_payout = net_commission_after_tax * (advisor_split_pct / 100.0)



    st.markdown("---")

    st.subheader("💰 Commission Breakdown & Tax Summary")



    m1, m2, m3, m4 = st.columns(4)

    m1.metric("Gross Commission", f"₱{gross_commission:,.2f}", f"{comm_pct:.2f}% of TCP")

    m2.metric("Less Withholding Tax", f"₱{tax_amount:,.2f}", f"-{tax_pct:.2f}% Tax Rate", delta_color="inverse")

    m3.metric("Net Commission (After Tax)", f"₱{net_commission_after_tax:,.2f}", f"Effective Rate: {net_effective_comm_rate:.3f}%")

    m4.metric("Your Net Payout Share", f"₱{advisor_payout:,.2f}", f"{advisor_split_pct:.2f}% Split Share")



    st.markdown("---")

    st.subheader("📋 Payout Record Summary")



    summary_text = f"""==================================================

            COMMISSION TRACKING SHEET

==================================================

CLIENT & PROPERTY:

• Borrower: {client_name}

• Project: {project_name}

• Total Contract Price: ₱{tcp_amount:,.2f}



COMMISSION & TAX DEDUCTIONS:

• Gross Commission Rate: {comm_pct:.2f}%

• Gross Commission Amount: ₱{gross_commission:,.2f}

• Less Withholding Tax ({tax_pct:.2f}%): -₱{tax_amount:,.2f}

• Net Commission (After Tax): ₱{net_commission_after_tax:,.2f}

• Effective Net Rate: {net_effective_comm_rate:.3f}% of TCP



ADVISOR PAYOUT:

• Advisor Split Share ({advisor_split_pct:.2f}%): ₱{advisor_payout:,.2f}



PAYOUT SCHEDULE:

• Trigger Milestone: {milestone}

• Target Release Date: {expected_date.strftime('%B %d, %Y')}

• Current Status: {status.upper()}

=================================================="""



    st.code(summary_text, language="text")



# --- MODE 13: CLIENT CALL & FOLLOW-UP SCHEDULER ---

elif app_mode == "Call & Follow-up Scheduler 📅":

    st.title("Client Call & Follow-up Scheduler 📅")

    st.write("Automatically prioritizes client outreach based on turnover due dates and past-due status.")



    default_sheet = "https://docs.google.com/spreadsheets/d/1TVrNsmb_J5RtMglqVNCtPcMD7vJKRiL9rGzoSGT2QRI/edit?gid=859880315#gid=859880315"

    sheet_input = st.text_input("Google Sheet Data Link:", value=default_sheet)



    if sheet_input:

        try:

            gid_part = ""

            if "gid=" in sheet_input:

                gid_string = sheet_input[sheet_input.find("gid="):]

                gid_value = gid_string.split('&')[0]

                gid_part = "&" + gid_value



            if "/edit" in sheet_input:

                csv_url = sheet_input.split("/edit")[0] + "/export?format=csv" + gid_part

            else:

                csv_url = sheet_input



            df_calls = load_cached_data(csv_url).copy()

            df_calls.columns = df_calls.columns.str.strip()



            name_col = None

            for c in ['PRINCIPAL BORROWERS NAME', 'NAME', 'BORROWER']:

                if c in df_calls.columns:

                    name_col = c

                    break



            if name_col and 'LSB DUE DATE' in df_calls.columns:

                today = pd.Timestamp.today().normalize()

                

                def assign_priority(due_str):

                    if pd.isna(due_str) or not str(due_str).strip():

                        return "Low Priority", 999, "No Date"

                    try:

                        due_dt = pd.to_datetime(str(due_str).strip())

                        diff_days = (due_dt - today).days

                        if diff_days < 0:

                            return "🔴 PAST DUE (Urgent)", diff_days, f"Past due by {abs(diff_days)} days"

                        elif diff_days <= 30:

                            return "🟠 HIGH (Due within 30 Days)", diff_days, f"Due in {diff_days} days"

                        elif diff_days <= 90:

                            return "🟡 MEDIUM (Due within 90 Days)", diff_days, f"Due in {diff_days} days"

                        else:

                            return "🟢 LOW (Future)", diff_days, f"Due in {diff_days} days"

                    except:

                        return "Low Priority", 999, "Invalid Date"



                df_calls[['Priority_Level', 'Days_Diff', 'Due_Status']] = df_calls['LSB DUE DATE'].apply(

                    lambda x: pd.Series(assign_priority(x))

                )



                if 'BANK STATUS' in df_calls.columns:

                    df_calls = df_calls[~df_calls['BANK STATUS'].astype(str).str.upper().isin(['CANCELLED', 'RELEASED'])]



                df_calls = df_calls.sort_values(by='Days_Diff')



                p_filter = st.multiselect(

                    "Filter Priority Levels:",

                    options=list(df_calls['Priority_Level'].unique()),

                    default=[p for p in df_calls['Priority_Level'].unique() if "🔴" in p or "🟠" in p]

                )



                if p_filter:

                    filtered_calls = df_calls[df_calls['Priority_Level'].isin(p_filter)]

                else:

                    filtered_calls = df_calls



                st.subheader(f"📋 Actionable Daily Call List ({len(filtered_calls)} Clients)")



                display_cols = [name_col, 'PROJECT', 'UNIT CODE', 'LUMPSUM BALANCE', 'LSB DUE DATE', 'Due_Status', 'Priority_Level']

                valid_cols = [c for c in display_cols if c in filtered_calls.columns]



                st.dataframe(filtered_calls[valid_cols], use_container_width=True)



        except Exception as e:

            st.error(f"Error parsing call schedule: {e}")
